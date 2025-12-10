"""
Excel hisobot yaratish moduli
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from datetime import datetime
from typing import List, Dict


def create_attendance_report(report_data: dict) -> BytesIO:
    """Haftalik davomat hisoboti Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Davomat"
    
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    present_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    absent_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    ws.merge_cells('A1:H1')
    ws['A1'] = "DAVOMAT HISOBOTI"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = center_align
    
    ws['A2'] = f"Yo'nalish: {report_data['direction']}"
    ws['A3'] = f"Guruh: {report_data['group_name']}"
    ws['A4'] = f"Davr: {report_data['week_start']} - {report_data['week_end']}"
    ws['A5'] = f"Yaratilgan: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    
    start_row = 7
    headers = ["№", "F.I.O", "ID"]
    sessions = report_data.get("sessions", [])
    
    for session in sessions:
        day_short = session["day"][:3]
        headers.append(f"{day_short}\n{session['date'][5:]}")
    
    headers.extend(["Keldi", "Kelmadi", "%"])
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    
    students = report_data.get("students", [])
    
    for idx, student in enumerate(students, 1):
        row = start_row + idx
        
        ws.cell(row=row, column=1, value=idx).border = thin_border
        ws.cell(row=row, column=1).alignment = center_align
        ws.cell(row=row, column=2, value=student["full_name"]).border = thin_border
        ws.cell(row=row, column=3, value=student.get("student_id", "")).border = thin_border
        ws.cell(row=row, column=3).alignment = center_align
        
        present_count = 0
        absent_count = 0
        
        for col_idx, session in enumerate(sessions, 4):
            status = student["attendance"].get(session["session_id"], "absent")
            cell = ws.cell(row=row, column=col_idx)
            
            if status == "present":
                cell.value = "+"
                cell.fill = present_fill
                present_count += 1
            else:
                cell.value = "-"
                cell.fill = absent_fill
                absent_count += 1
            
            cell.alignment = center_align
            cell.border = thin_border
        
        stat_col = len(sessions) + 4
        
        ws.cell(row=row, column=stat_col, value=present_count).border = thin_border
        ws.cell(row=row, column=stat_col).alignment = center_align
        ws.cell(row=row, column=stat_col).fill = present_fill
        
        ws.cell(row=row, column=stat_col + 1, value=absent_count).border = thin_border
        ws.cell(row=row, column=stat_col + 1).alignment = center_align
        ws.cell(row=row, column=stat_col + 1).fill = absent_fill
        
        total = present_count + absent_count
        percentage = (present_count / total * 100) if total > 0 else 0
        pct_cell = ws.cell(row=row, column=stat_col + 2, value=f"{percentage:.0f}%")
        pct_cell.border = thin_border
        pct_cell.alignment = center_align
        
        if percentage >= 80:
            pct_cell.fill = present_fill
        elif percentage < 50:
            pct_cell.fill = absent_fill
    
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 12
    
    for col in range(4, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 10
    
    ws.row_dimensions[start_row].height = 35
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def create_student_report(student_data: dict, attendance_list: List[Dict]) -> BytesIO:
    """Talabaning shaxsiy davomat hisoboti"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Shaxsiy davomat"
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
    present_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    absent_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    ws.merge_cells('A1:E1')
    ws['A1'] = "SHAXSIY DAVOMAT HISOBOTI"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = center_align
    
    ws['A3'] = "F.I.O:"
    ws['B3'] = student_data.get("full_name", "")
    ws['A4'] = "ID:"
    ws['B4'] = student_data.get("student_id", "")
    ws['A5'] = "Yo'nalish:"
    ws['B5'] = student_data.get("direction", "")
    ws['A6'] = "Guruh:"
    ws['B6'] = student_data.get("group_name", "")
    
    start_row = 8
    headers = ["№", "Fan", "Sana", "Kun", "Holat"]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    
    present_count = 0
    absent_count = 0
    
    for idx, record in enumerate(attendance_list, 1):
        row = start_row + idx
        
        ws.cell(row=row, column=1, value=idx).border = thin_border
        ws.cell(row=row, column=1).alignment = center_align
        ws.cell(row=row, column=2, value=record["subject"]).border = thin_border
        ws.cell(row=row, column=3, value=record["date"]).border = thin_border
        ws.cell(row=row, column=3).alignment = center_align
        ws.cell(row=row, column=4, value=record["day"]).border = thin_border
        ws.cell(row=row, column=4).alignment = center_align
        
        status_cell = ws.cell(row=row, column=5)
        if record["status"] == "present":
            status_cell.value = "Keldi"
            status_cell.fill = present_fill
            present_count += 1
        else:
            status_cell.value = "Kelmadi"
            status_cell.fill = absent_fill
            absent_count += 1
        status_cell.border = thin_border
        status_cell.alignment = center_align
    
    stat_row = start_row + len(attendance_list) + 2
    ws.cell(row=stat_row, column=1, value="JAMI:")
    ws.cell(row=stat_row, column=1).font = Font(bold=True)
    
    total = present_count + absent_count
    percentage = (present_count / total * 100) if total > 0 else 0
    
    ws.cell(row=stat_row + 1, column=1, value=f"Keldi: {present_count}")
    ws.cell(row=stat_row + 2, column=1, value=f"Kelmadi: {absent_count}")
    ws.cell(row=stat_row + 3, column=1, value=f"Davomat: {percentage:.1f}%")
    ws.cell(row=stat_row + 3, column=1).font = Font(bold=True)
    
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 15
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
